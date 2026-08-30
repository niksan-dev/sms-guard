"""
Template-based Site Invoice generation.

The Excel workbook is the source of truth for the visual layout.
Python only fills mapped cells and converts the finished workbook to PDF.
"""

from __future__ import annotations
import time
import io
import os
import shutil
import subprocess
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from sqlalchemy.orm import joinedload


from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

from database.connection import SessionLocal
from database.site_bill import SiteBill
from database.company_settings import CompanySettings
from database.models import Site

# Project root: <project>/services/site_invoice_template_service.py
PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = PROJECT_ROOT / "templates" / "site_invoice"
DEFAULT_TEMPLATE = TEMPLATE_DIR / "template_1.xlsx"


# ============================================================
# TEMPLATE DISCOVERY
# ============================================================

def list_site_invoice_templates() -> dict[str, Path]:
    """Return available site invoice templates."""

    if not TEMPLATE_DIR.exists():
        return {}

    templates: dict[str, Path] = {}

    for path in sorted(TEMPLATE_DIR.glob("*.xlsx")):
        templates[path.stem] = path

    return templates


def get_default_site_invoice_template() -> Path:
    """Return the default Aadhar invoice template."""

    if not DEFAULT_TEMPLATE.exists():
        raise FileNotFoundError(
            f"Site invoice template not found: {DEFAULT_TEMPLATE}"
        )

    return DEFAULT_TEMPLATE


# ============================================================
# HELPERS
# ============================================================

def _safe(value: Any, default: str = "") -> str:
    if value is None:
        return default
    return str(value)


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        value = value.date()

    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    return _safe(value)


def _month_text(year: int, month: int) -> str:
    return date(int(year), int(month), 1).strftime("%b %y").upper()


def _amount_words(number: float) -> str:
    """Indian-friendly English amount in words, with paise."""

    number = round(float(number or 0), 2)
    rupees = int(number)
    paise = int(round((number - rupees) * 100))

    ones = [
        "Zero", "One", "Two", "Three", "Four", "Five", "Six",
        "Seven", "Eight", "Nine", "Ten", "Eleven", "Twelve",
        "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen",
        "Eighteen", "Nineteen",
    ]

    tens = [
        "", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty",
        "Seventy", "Eighty", "Ninety",
    ]

    def under_100(n: int) -> str:
        if n < 20:
            return ones[n]

        return tens[n // 10] + (
            f" {ones[n % 10]}"
            if n % 10
            else ""
        )

    def under_1000(n: int) -> str:
        if n < 100:
            return under_100(n)

        rest = n % 100

        result = f"{ones[n // 100]} Hundred"

        if rest:
            result += f" {under_100(rest)}"

        return result

    def indian_integer(n: int) -> str:
        if n == 0:
            return "Zero"

        parts: list[str] = []

        crore = n // 10_000_000
        n %= 10_000_000

        lakh = n // 100_000
        n %= 100_000

        thousand = n // 1_000
        n %= 1_000

        if crore:
            parts.append(
                f"{under_1000(crore)} Crore"
            )

        if lakh:
            parts.append(
                f"{under_100(lakh)} Lakh"
            )

        if thousand:
            parts.append(
                f"{under_100(thousand)} Thousand"
            )

        if n:
            parts.append(
                under_1000(n)
            )

        return " ".join(parts)

    result = indian_integer(rupees) + " Rupees"

    if paise:
        result += (
            f" and {under_100(paise)} Paise"
        )

    return result + " Only."


def _get_attr(
    obj: Any,
    name: str,
    default: Any = None
) -> Any:
    """Read an already-loaded Python/SQLAlchemy attribute safely."""

    if obj is None:
        return default

    return getattr(
        obj,
        name,
        default
    )


def _client_name(client: Any) -> str:
    if client is None:
        return ""

    return (
        _get_attr(client, "username")
        or _get_attr(client, "name")
        or _get_attr(client, "company_name")
        or ""
    )


def _client_pan(client: Any) -> str:
    """Return client PAN if the User model eventually provides one."""

    if client is None:
        return ""

    return (
        _get_attr(client, "pan_number")
        or _get_attr(client, "pan")
        or ""
    )


def _site_address(site: Any) -> str:
    parts = [
        _get_attr(site, "address", ""),
        _get_attr(site, "city", ""),
        _get_attr(site, "state", ""),
        _get_attr(site, "pincode", ""),
    ]

    return ", ".join(
        str(part).strip()
        for part in parts
        if part and str(part).strip()
    )


class CompanySettingsRequiredError(RuntimeError):
    """Raised when an invoice is requested before company settings exist."""


def _company_values() -> dict[str, Any]:
    """
    Return a detached snapshot of the configured company settings.

    The invoice generator deliberately queries CompanySettings directly so
    the PDF never depends on a detached SQLAlchemy object.
    """

    db = SessionLocal()

    try:
        settings = (
            db.query(CompanySettings)
            .order_by(
                CompanySettings.id.asc()
            )
            .first()
        )

        if settings is None:
            raise CompanySettingsRequiredError(
                "Company settings are not configured. "
                "Please configure Company Settings before generating an invoice."
            )

        return {
            "id": settings.id,
            "company_name": settings.company_name,
            "owner_name": settings.owner_name,
            "phone": settings.phone,
            "email": settings.email,
            "address": settings.address,
            "city": settings.city,
            "state": settings.state,
            "pincode": settings.pincode,
            "gst_number": settings.gst_number,
            "pan_number": settings.pan_number,
            "bank_name": settings.bank_name,
            "account_holder_name": settings.account_holder_name,
            "account_number": settings.account_number,
            "ifsc_code": settings.ifsc_code,
            "branch_name": settings.branch_name,
            "invoice_prefix": settings.invoice_prefix,
            "gst_invoice_prefix": settings.gst_invoice_prefix,
            "logo_path": settings.logo_path,
            "updated_at": settings.updated_at,
            "cgst_rate":settings.cgst_rate,
            "sgst_rate":settings.sgst_rate,
            "alternate_phone":settings.alternate_phone,
        }

    finally:
        db.close()


def _resolve_logo_path(
    logo_path: Any
) -> Path | None:
    """Resolve CompanySettings.logo_path against the project root."""

    if not logo_path:
        return None

    path = Path(
        str(logo_path)
    )

    if not path.is_absolute():
        path = PROJECT_ROOT / path

    try:
        path = path.resolve()
    except OSError:
        return None

    if path.is_file():
        return path

    return None


def _apply_company_logo(
    ws: Any,
    settings: dict[str, Any]
) -> None:
    """Replace the template logo with the logo configured in Company Settings."""

    # Remove the static/template logo so a stale logo can never be shown.
    ws._images = []

   # print("logo_path=============>>>>>",settings.get("logo_path"))

    logo_path = _resolve_logo_path(
        settings.get("logo_path")
    )

    if logo_path is None:
        return

    try:
        image = XLImage(
            str(logo_path)
        )
    except Exception as exc:
        raise ValueError(
            f"Unable to load company logo: "
            f"{logo_path} ({exc})"
        ) from exc

    image.width = 110
    image.height = 110
    image.anchor = get_image_anchor(ws,image)

    

    ws.add_image(image)

def get_image_anchor(ws,image):
    # Merged range: A2:C5
    # Calculate approximate width of A:C
    total_width = sum(
        ws.column_dimensions[col].width or 8.43
        for col in ["A", "B"]
    )

    # Excel column width -> approximate pixels
    total_width_px = int(total_width * 7)

    # Calculate row height in pixels
    total_height = sum(
        ws.row_dimensions[row].height or 15
        for row in range(2, 5)
    )

    # Excel points -> pixels
    total_height_px = int(total_height * 96 / 72)

    # Center position
    offset_x = max(0, (total_width_px - image.width) // 2)
    offset_y = max(0, (total_height_px - image.height) // 2)

    # Create anchor
    anchor = OneCellAnchor(
        _from=ws["A2"]._from if hasattr(ws["A2"], "_from") else None,
        ext=XDRPositiveSize2D(
            cx=pixels_to_EMU(image.width),
            cy=pixels_to_EMU(image.height)
        )
    )

    # Set starting cell
    anchor._from.col = 0       # A
    anchor._from.row = 1       # row 2

    # Offset from A2
    anchor._from.colOff = pixels_to_EMU(offset_x)
    anchor._from.rowOff = pixels_to_EMU(offset_y)

    return anchor
# ============================================================
# DATABASE SNAPSHOT
# ============================================================

def _load_bill_snapshot(
    bill: Any
) -> dict[str, Any]:
    """
    Reload the SiteBill and its Site/Client relationships while the
    SQLAlchemy session is open, then convert everything needed by the
    template into ordinary Python values.

    This deliberately prevents DetachedInstanceError when the caller
    passes a SiteBill whose original session has already been closed.
    """

    bill_id = _get_attr(
        bill,
        "id"
    )

    if bill_id is None:
        raise ValueError(
            "Site bill ID is required."
        )

    db = SessionLocal()

    try:
        loaded_bill = (
            db.query(SiteBill)
            .options(
                joinedload(
                    SiteBill.site
                ).joinedload(
                    Site.client
                )
            )
            .filter(
                SiteBill.id == int(bill_id)
            )
            .first()
        )

        if loaded_bill is None:
            raise ValueError(
                f"Site bill not found: {bill_id}"
            )

        site = loaded_bill.site
        client = (
            site.client
            if site is not None
            else None
        )

        # Read all relationship/scalar values before closing the session.
        return {
            "bill_number": loaded_bill.bill_number,
            "billing_month": loaded_bill.billing_month,
            "billing_year": loaded_bill.billing_year,
            "bill_date": loaded_bill.bill_date,
            "total_days": loaded_bill.total_days,
            "shift_1_count": loaded_bill.shift_1_count,
            "shift_2_count": loaded_bill.shift_2_count,
            "total_shifts": loaded_bill.total_shifts,
            "monthly_rate": loaded_bill.monthly_rate,
            "shift_rate": loaded_bill.shift_rate,
            "gross_amount": loaded_bill.gross_amount,
            "cgst_amount": loaded_bill.cgst_amount,
            "sgst_amount": loaded_bill.sgst_amount,
            "total_amount": loaded_bill.total_amount,
            "site_code": (
                _get_attr(
                    site,
                    "site_code",
                    ""
                )
                if site
                else ""
            ),
            "site_name": (
                _get_attr(
                    site,
                    "name",
                    ""
                )
                if site
                else ""
            ),
            "site_email": (
                _get_attr(
                    site,
                    "email",
                    ""
                )
                if site
                else ""
            ),
            "site_address": (
                _site_address(site)
                if site
                else ""
            ),
            "customer_name": _client_name(client),
            "customer_email": (
                _get_attr(
                    client,
                    "email",
                    ""
                )
                if client
                else ""
            ),
            "customer_pan": _client_pan(client),
            "updated_at": loaded_bill.updated_at,
        }

    finally:
        db.close()


# ============================================================
# BUILD EXCEL FROM TEMPLATE
# ============================================================

def build_site_invoice_workbook(
    bill: Any,
    template_path: str | os.PathLike | None = None,
) -> bytes:
    """
    Fill the supplied Excel invoice template and return XLSX bytes.

    Reference invoice mapping:

        C2       Company name
        C3       Company address
        C4       Company phone / email

        A5       Bill From + company name ONLY
        D5       Bill To + customer name + site name
        H5       Invoice number
        H7       Invoice date

        A9       Company PAN

        A10      Site address

        B13      Billing month
        B15:F15  Shift 1
        B16:F16  Shift 2

        F20      Gross amount
        F21      CGST
        F22      SGST
        A23      Amount in words
        G23      Grand Total
        A24      Bank details
        E24      Authorized signature

    Important layout change:
        - Company PAN is NOT written inside Bill From (A5).
        - Company PAN is written in the row below Bill From (A9).
        - Site name is shown below the customer name in Bill To (D5).
    """

    path = (
        Path(template_path)
        if template_path
        else get_default_site_invoice_template()
    )

    if not path.exists():
        raise FileNotFoundError(
            f"Invoice template not found: {path}"
        )

    workbook = load_workbook(path)

    if "Invoice" not in workbook.sheetnames:
        raise ValueError(
            "The selected invoice template must contain "
            "an 'Invoice' sheet."
        )

    ws = workbook["Invoice"]

    if "Python Mapping" in workbook.sheetnames:
        workbook["Python Mapping"].sheet_state = "hidden"

    snapshot = _load_bill_snapshot(
        bill
    )

    settings = _company_values()

   # print("CGST RATE----->>>>",settings.get("cgst_rate"))

    billing_year = int(
        snapshot["billing_year"]
        or date.today().year
    )

    billing_month = int(
        snapshot["billing_month"]
        or date.today().month
    )

    invoice_date = (
        snapshot["bill_date"]
        or date.today()
    )

    invoice_number = snapshot[
        "bill_number"
    ]

    # ========================================================
    # COMPANY SETTINGS
    # ========================================================

    company_name = _safe(
        settings.get("company_name")
    )

    company_address = _safe(
        settings.get("address")
    )

    company_city = _safe(
        settings.get("city")
    )

    company_state = _safe(
        settings.get("state")
    )

    company_pincode = _safe(
        settings.get("pincode")
    )

    company_phone = _safe(
        settings.get("phone")
    )

    company_alternate_phone = _safe(
            settings.get("alternate_phone")
        )

    company_email = _safe(
        settings.get("email")
    )

    company_pan = _safe(
        settings.get("pan_number")
    )

    company_gst = _safe(
        settings.get("gst_number")
    )

    owner_name = _safe(
        settings.get("owner_name")
    )

    bank_name = _safe(
        settings.get("bank_name")
    )

    account_holder_name = _safe(
        settings.get("account_holder_name")
    )

    account_number = _safe(
        settings.get("account_number")
    )

    ifsc_code = _safe(
        settings.get("ifsc_code")
    )

    branch_name = _safe(
        settings.get("branch_name")
    )

    full_company_address = ", ".join(
        str(part).strip()
        for part in [
            company_address,
            company_city,
            company_state,
            company_pincode,
        ]
        if part and str(part).strip()
    )

    cgst_rate = _safe(settings.get("cgst_rate"))

    sgst_rate = _safe(settings.get("sgst_rate"))


    # ========================================================
    # HEADER
    # ========================================================

    _apply_company_logo(
        ws,
        settings
    )

    ws["C2"] = company_name

    ws["C3"] = (
        full_company_address
        if full_company_address
        else ""
    )

    contact_parts = []

    if company_phone:
        contact_parts.append(
            company_phone
        )
    if company_alternate_phone:
        contact_parts.append(
            company_alternate_phone
        )

    # if company_email:
    #     contact_parts.append(
    #         company_email
    #     )

    ws["C4"] = " / ".join(
        contact_parts
    )

    # ========================================================
    # BILL FROM
    #
    # PAN deliberately removed from this cell.
    # ========================================================

    ws["A5"] = (
        f"Bill From,\n"
        f"{company_name}"
    )

    # Optional GST can remain in the dedicated lower identity
    # row only if the template has a place for it. For the supplied
    # reference layout, only PAN is placed here.
    #
    # Company PAN is written separately below Bill From.
    ws["A9"] = (
        f"PAN:- {company_pan}"
        if company_pan
        else "PAN:- "
    )

    # ========================================================
    # BILL TO
    #
    # Customer name + Site name.
    # ========================================================

    customer_name = _safe(
        snapshot["customer_name"]
    )

    site_name = _safe(
        snapshot["site_name"]
    )

    bill_to_lines = [
        "Bill To,"
    ]

    if customer_name:
        bill_to_lines.append(
            customer_name
        )

    if site_name:
        bill_to_lines.append(
            site_name
        )

    ws["D5"] = "\n".join(
        bill_to_lines
    )

    # ========================================================
    # INVOICE NUMBER / DATE
    # ========================================================

    ws["H5"] = invoice_number

    ws["H7"] = _date_text(
        invoice_date
    )

    # ========================================================
    # CLIENT PAN
    #
    # User model currently may not have PAN. If present later,
    # use it here. Never substitute company PAN.
    # ========================================================

    customer_pan = _safe(
        snapshot["customer_pan"]
    )

    ws["D9"] = (
        f"PAN:- {customer_pan}"
        if customer_pan
        else "PAN:- "
    )

    # ========================================================
    # SITE ADDRESS
    # ========================================================

    site_address = _safe(
        snapshot["site_address"]
    )

    ws["A10"] = (
        f"Site Address:- {site_address}"
        if site_address
        else "Site Address:-"
    )

    # ========================================================
    # BILLING DATA
    # ========================================================

    days = int(
        snapshot["total_days"]
        or 0
    )

    shift_rate = _number(
        snapshot["shift_rate"]
    )

    shift_1_count = int(
        snapshot["shift_1_count"]
        or 0
    )

    shift_2_count = int(
        snapshot["shift_2_count"]
        or 0
    )

    shift_1_amount = round(
        shift_rate * shift_1_count,
        2
    )

    shift_2_amount = round(
        shift_rate * shift_2_count,
        2
    )

    ws["B13"] = (
        f"For the month of\n"
        f"{_month_text(billing_year, billing_month)}"
    )

    # Shift 1
    ws["B15"] = "Day/Night"
    ws["C15"] = "01"
    ws["D15"] = shift_1_count
    ws["E15"] = days
    ws["F15"] = shift_1_amount

    # Shift 2
    ws["B16"] = "Day/Night"
    ws["C16"] = "01"
    ws["D16"] = shift_2_count
    ws["E16"] = days
    ws["F16"] = shift_2_amount

    gross_amount = _number(
        snapshot["gross_amount"]
    )

    cgst_amount = _number(
        snapshot["cgst_amount"]
    )

    sgst_amount = _number(
        snapshot["sgst_amount"]
    )

    total_amount = _number(
        snapshot["total_amount"]
    )

    # ========================================================
    # TOTALS
    # ========================================================

    ws["F20"] = gross_amount

    ws["F21"] = (
        cgst_amount
        if cgst_amount
        else "0"
    )

    ws["A21"] = (
        "Add CGST @ "
        f"{cgst_rate} of Gross"
    )

    ws["A22"] = (
            "Add SGST @ "
            f"{sgst_rate} of Gross"
        )

    ws["F22"] = (
        sgst_amount
        if sgst_amount
        else "0"
    )

    ws["A23"] = (
        "Total Amt In Word :- "
        f"{_amount_words(total_amount)}"
    )

    ws["G23"] = (total_amount)

    # ========================================================
    # BANK DETAILS
    # ========================================================

    bank_lines = [
        "BANK DETAILS:-"
    ]

    if bank_name:
        bank_lines.append(
            f"Bank name :-"
            f"{bank_name}"
        )
    elif company_name:
        bank_lines.append(
            company_name
        )

    if account_holder_name:
        bank_lines.append(
            f"A/C HOLDER :- "
            f"{account_holder_name}"
        )

    if account_number:
        bank_lines.append(
            f"ACC NO :- "
            f"{account_number}"
        )

    if ifsc_code:
        bank_lines.append(
            f"IFSC CODE:- "
            f"{ifsc_code}"
        )

    if branch_name:
        bank_lines.append(
            f"BRANCH:- "
            f"{branch_name}."
        )

    ws["A24"] = "\n".join(
        bank_lines
    )

    # ========================================================
    # SIGNATURE
    # ========================================================

    signature_name = (
        owner_name
        or "Proprietor"
    )

    ws["E24"] = (
        f"FOR {company_name}.\n\n"
        f"Proprietor"
    )

    # ========================================================
    # PAGE / PRINT SETTINGS
    # ========================================================

    ws.print_area = "A1:H38"

    ws.page_setup.paperSize = (
        ws.PAPERSIZE_A4
    )

    ws.page_setup.orientation = (
        "portrait"
    )

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.20
    ws.page_margins.right = 0.20
    ws.page_margins.top = 0.20
    ws.page_margins.bottom = 0.20

    # ========================================================
    # FORMATTING
    # ========================================================

    for cell in (
        "F15",
        "F16",
        "F20",
        "F21",
        "F22",
    ):
        ws[cell].number_format = (
            '#,##0.00'
        )

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"

    output = io.BytesIO()

    workbook.save(
        output
    )

    return output.getvalue()


# ============================================================
# PDF CACHE
# ============================================================

# Streamlit reruns the script whenever a button is clicked.  Keep generated
# PDFs in this process so View -> Export -> Mail does not start Excel again.
_PDF_CACHE: dict[tuple[Any, ...], bytes] = {}


def _pdf_cache_key(
    bill: Any,
    template_path: str | os.PathLike | None,
) -> tuple[Any, ...]:
    """
    Build a cache key that invalidates whenever the bill,
    company settings, or template changes.
    """

    # --------------------------------------------------------
    # IMPORTANT:
    # Always reload the bill from DB.
    #
    # Streamlit may rerun with an older SiteBill Python object,
    # so we must NOT rely only on bill.updated_at here.
    # --------------------------------------------------------

    snapshot = _load_bill_snapshot(bill)

    bill_key = (
        snapshot.get("bill_number"),
        snapshot.get("billing_year"),
        snapshot.get("billing_month"),
        snapshot.get("bill_date"),

        snapshot.get("total_days"),
        snapshot.get("shift_1_count"),
        snapshot.get("shift_2_count"),
        snapshot.get("total_shifts"),

        snapshot.get("monthly_rate"),
        snapshot.get("shift_rate"),

        # IMPORTANT
        snapshot.get("gross_amount"),
        snapshot.get("cgst_amount"),
        snapshot.get("sgst_amount"),
        snapshot.get("total_amount"),

        # Database modification timestamp
        snapshot.get("updated_at"),
    )

    # --------------------------------------------------------
    # TEMPLATE VERSION
    # --------------------------------------------------------

    try:

        template = (
            Path(template_path)
            if template_path
            else get_default_site_invoice_template()
        )

        template_key = (
            str(template.resolve()),
            template.stat().st_mtime_ns,
            template.stat().st_size,
        )

    except (OSError, ValueError):

        template_key = (
            str(
                template_path
                or DEFAULT_TEMPLATE
            ),
        )

    # --------------------------------------------------------
    # COMPANY SETTINGS VERSION
    # --------------------------------------------------------

    settings = _company_values()

    company_key = (
        settings.get("id"),
        settings.get("updated_at"),
        settings.get("logo_path"),
    )

    # --------------------------------------------------------
    # FINAL CACHE KEY
    # --------------------------------------------------------

    return (
        bill_key,
        company_key,
        template_key,
    )


def clear_site_invoice_pdf_cache() -> None:
    """Clear all cached site invoice PDFs."""
    _PDF_CACHE.clear()


# ============================================================
# CONVERT XLSX -> PDF
# ============================================================

def _find_soffice():
    candidates = [
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        "/usr/bin/soffice",
        "/usr/bin/libreoffice",
    ]

    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(candidate)

    return None


def _excel_com_to_pdf(
    xlsx_data: bytes
) -> bytes | None:
    """Convert XLSX to PDF with Microsoft Excel COM.

    Streamlit can execute the same Python process across many reruns.
    Explicit COM initialization and cleanup prevents Excel COM from becoming
    unavailable after the first conversion.
    """

    if os.name != "nt":
        return None

    try:
        import pythoncom  # type: ignore
        import win32com.client  # type: ignore
    except ImportError:
        return None

    excel = None
    workbook = None
    com_initialized = False

    with tempfile.TemporaryDirectory(
        prefix="site_invoice_excel_"
    ) as temp_dir:

        temp_path = Path(temp_dir)
        xlsx_path = temp_path / "site_invoice.xlsx"
        pdf_path = temp_path / "site_invoice.pdf"

        xlsx_path.write_bytes(xlsx_data)

        try:
            # Important for Streamlit/Windows: initialize COM in the current
            # worker thread before touching Excel.
            pythoncom.CoInitialize()
            com_initialized = True

            excel = win32com.client.DispatchEx(
                "Excel.Application"
            )

            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            excel.EnableEvents = False

            workbook = excel.Workbooks.Open(
                str(xlsx_path),
                ReadOnly=True,
                UpdateLinks=0,
                IgnoreReadOnlyRecommended=True,
                AddToMru=False,
            )

            # 0 = xlTypePDF, 0 = xlQualityStandard
            workbook.ExportAsFixedFormat(
                0,
                str(pdf_path),
                0,
                True,
                False,
            )

            if pdf_path.exists() and pdf_path.stat().st_size > 0:
                return pdf_path.read_bytes()

        except Exception:
            # Excel is optional; caller can fall back to LibreOffice.
            return None

        finally:
            if workbook is not None:
                try:
                    workbook.Close(False)
                except Exception:
                    pass

            if excel is not None:
                try:
                    excel.DisplayAlerts = False
                except Exception:
                    pass

                try:
                    excel.Quit()
                except Exception:
                    pass

            if com_initialized:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    return None


def xlsx_bytes_to_pdf(
    xlsx_data: bytes
) -> bytes:
    """
    Convert an XLSX byte stream to PDF.

    This function is intentionally kept conversion-only. Caching is performed
    by generate_site_bill_pdf_from_template(), where the bill/template/company
    identity is available.
    """

    excel_pdf = _excel_com_to_pdf(
        xlsx_data
    )

    if excel_pdf:
        return excel_pdf

    soffice = _find_soffice()

    if not soffice:
        raise RuntimeError(
            "Neither Microsoft Excel nor LibreOffice is available "
            "for Excel-to-PDF conversion. Install LibreOffice, or "
            "install pywin32 with Microsoft Excel on Windows."
        )

    with tempfile.TemporaryDirectory(
        prefix="site_invoice_"
    ) as temp_dir:

        temp_path = Path(temp_dir)
        xlsx_path = temp_path / "site_invoice.xlsx"
        xlsx_path.write_bytes(xlsx_data)

        profile_dir = temp_path / "lo_profile"
        profile_dir.mkdir()

        command = [
            soffice,
            "--headless",
            (
                "-env:UserInstallation="
                f"file:///{profile_dir.as_posix()}"
            ),
            "--convert-to",
            "pdf",
            "--outdir",
            str(temp_path),
            str(xlsx_path),
        ]

        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            timeout=60,
        )

        pdf_path = temp_path / "site_invoice.pdf"

        if (
            result.returncode != 0
            or not pdf_path.exists()
            or pdf_path.stat().st_size == 0
        ):
            raise RuntimeError(
                "Excel template PDF conversion failed. "
                f"stdout={result.stdout.strip()} "
                f"stderr={result.stderr.strip()}"
            )

        return pdf_path.read_bytes()


# ============================================================
# PUBLIC PDF API
# ============================================================

def generate_site_bill_pdf_from_template(
    bill: Any,
    template_path: str | os.PathLike | None = None,
) -> bytes:

    total_start = time.perf_counter()

    print("\n========================================")
    print("START PDF GENERATION")
    print("========================================")

    start = time.perf_counter()

    cache_key = _pdf_cache_key(
        bill,
        template_path,
    )

    print(
        f"[1] Cache key: "
        f"{time.perf_counter() - start:.3f}s"
    )

    cached_pdf = _PDF_CACHE.get(
        cache_key
    )

    if cached_pdf is not None:

        print(
            f"[CACHE HIT] "
            f"{time.perf_counter() - total_start:.3f}s"
        )

        return cached_pdf

    start = time.perf_counter()

    xlsx_data = build_site_invoice_workbook(
        bill=bill,
        template_path=template_path,
    )

    print(
        f"[2] Build XLSX: "
        f"{time.perf_counter() - start:.3f}s"
    )

    start = time.perf_counter()

    pdf_data = xlsx_bytes_to_pdf(
        xlsx_data
    )

    print(
        f"[3] XLSX → PDF: "
        f"{time.perf_counter() - start:.3f}s"
    )

    _PDF_CACHE[cache_key] = pdf_data

    print(
        f"[4] TOTAL: "
        f"{time.perf_counter() - total_start:.3f}s"
    )

    print("========================================\n")

    return pdf_data

